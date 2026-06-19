"""
확인 체크리스트 엑셀 생성 스크립트
실행: python scripts/make_checklist_excel.py
생성 파일: output/전임자확인_체크리스트.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "output" / "전임자확인_체크리스트.xlsx"
OUTPUT_PATH.parent.mkdir(exist_ok=True)

# 색상 정의
COLOR_HEADER   = "1F4E79"   # 진한 파랑 (제목)
COLOR_DONE     = "E2EFDA"   # 연두 (확인 완료)
COLOR_TODO     = "FFF2CC"   # 연노랑 (확인 필요)
COLOR_SECTION  = "D6E4F0"   # 연파랑 (섹션 제목)
COLOR_ANSWER   = "FFFFFF"   # 흰색 (답변 입력칸)
COLOR_SUMMARY  = "F2F2F2"   # 연회색 (요약표)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(white=False):
    return Font(name="맑은 고딕", bold=True, color="FFFFFF" if white else "1F4E79", size=11)

def normal_font(bold=False, color="000000", size=10):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def set_cell(ws, row, col, value, font=None, fill_color=None, align="left", wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill_color:
        cell.fill = fill(fill_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = thin_border()
    return cell

def merge(ws, r1, c1, r2, c2, value, font=None, fill_color=None, align="left", wrap=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:
        cell.font = font
    if fill_color:
        cell.fill = fill(fill_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell

def build_checklist():
    wb = Workbook()
    ws = wb.active
    ws.title = "확인 체크리스트"

    # 열 너비 설정
    ws.column_dimensions["A"].width = 4    # 번호/체크
    ws.column_dimensions["B"].width = 38   # 질문
    ws.column_dimensions["C"].width = 40   # 답변
    ws.column_dimensions["D"].width = 14   # 확인자
    ws.column_dimensions["E"].width = 14   # 확인 일자

    row = 1

    # ── 타이틀 ──────────────────────────────────────────
    merge(ws, row, 1, row, 5,
          "근무표 자동화 프로그램 — 전임자 / 팀장 확인 체크리스트",
          font=Font(name="맑은 고딕", bold=True, color="FFFFFF", size=14),
          fill_color=COLOR_HEADER, align="center")
    ws.row_dimensions[row].height = 36
    row += 1

    merge(ws, row, 1, row, 5,
          "아래 항목을 확인하신 후 [답변] 칸에 내용을 입력해 주세요. 확인 완료 시 ✅ 체크 부탁드립니다.",
          font=normal_font(color="595959"),
          fill_color="F7F7F7", align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # ── 헤더 행 ─────────────────────────────────────────
    headers = ["✅", "질문 항목", "답변 (직접 입력)", "확인자", "확인 일자"]
    for col, h in enumerate(headers, 1):
        set_cell(ws, row, col, h,
                 font=header_font(white=True),
                 fill_color=COLOR_HEADER,
                 align="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # ── 확인 완료 항목 ────────────────────────────────────
    merge(ws, row, 1, row, 5,
          "✅  확인 완료 항목 (엑셀 분석으로 이미 파악됨)",
          font=normal_font(bold=True, color="375623"),
          fill_color=COLOR_DONE, align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    done_items = [
        ("근무 코드",    "빈칸(출근) / OFF(휴무) 두 가지만 사용"),
        ("파출 코드",    "-(불필요) / 1(1명 투입) 두 가지"),
        ("조 구성",      "주간조 6명 + 야간조 6명 (조리사 별도)"),
        ("근무 시간",    "주간/야간 고정시간 출퇴근 — 별도 기록 불필요"),
        ("파출 집계",    "월말 파출 필요 총 횟수 합산 자동 표기"),
    ]
    for label, content in done_items:
        set_cell(ws, row, 1, "✅", font=normal_font(bold=True, color="375623"), fill_color=COLOR_DONE, align="center")
        set_cell(ws, row, 2, f"[{label}]  {content}", font=normal_font(color="375623"), fill_color=COLOR_DONE)
        merge(ws, row, 3, row, 5, "", fill_color=COLOR_DONE)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # 빈 줄

    # ── 확인 필요 항목 ────────────────────────────────────
    sections = [
        {
            "no": 1,
            "title": "동시 OFF 제한",
            "questions": [
                "같은 조에서 하루에 최대 몇 명까지 동시에 OFF 가능한가요?\n  예) 1명만 가능 / 2명까지 가능(파출 보완) / 제한 없음",
            ]
        },
        {
            "no": 2,
            "title": "월 OFF 횟수 기준",
            "questions": [
                "직원 1인당 한 달에 OFF를 몇 번 주어야 하나요?\n  예) 매주 1회 / 고정 8회 / 근무일의 일정 비율",
                "주휴일 / 무휴 / 공휴 구분 방법은 어떻게 되나요?\n  예) 주 1회 보장이 주휴, 추가 휴무는 무휴",
            ]
        },
        {
            "no": 3,
            "title": "최소 유지 인원",
            "questions": [
                "평일 최소 몇 명이 근무해야 하나요? (엑셀 분석: 5명 추정)",
                "일요일은 몇 명까지 줄어도 되나요? (엑셀 분석: 4명 추정)",
                "토요일도 일요일과 같은 기준인가요?",
                "공휴일은 어떻게 처리하나요? (평일 기준 / 일요일 기준 / 별도 기준)",
            ]
        },
        {
            "no": 4,
            "title": "파출 투입 조건",
            "questions": [
                "최소 인원 미달 시 무조건 파출을 투입하나요?",
                "일요일에 4명이 되어도 파출 없이 운영하는 경우가 있나요? (엑셀에서 발견됨)",
                "파출 투입이 불가능한 경우는 어떻게 처리하나요?",
            ]
        },
        {
            "no": 5,
            "title": "조리사(반장) 스케줄",
            "questions": [
                "조리사(반장)는 시간제 직원과 같은 표에 관리하나요?",
                "조리사도 OFF가 있나요? (엑셀에서는 모두 빈칸으로 확인됨)",
                "조리사 OFF가 있다면 별도로 관리하나요?",
            ]
        },
        {
            "no": 6,
            "title": "OFF 신청 및 우선순위",
            "questions": [
                "직원이 특정 날짜 OFF를 신청하면 무조건 반영해야 하나요?",
                "신청이 겹칠 때 (2명이 같은 날 신청) 처리 기준은?\n  예) 선착순 / 조장이 조율 / 담당자가 결정",
                "OFF 신청 마감일이 있나요? (예: 전월 20일까지)",
            ]
        },
        {
            "no": 7,
            "title": "연차 처리",
            "questions": [
                "연차 사용 시 OFF와 동일하게 처리하나요?",
                "연차는 근무표에 별도로 표시해야 하나요?",
                "엑셀에서 연차 열이 모두 0인데, 실제 연차 사용자가 있나요?",
            ]
        },
        {
            "no": 8,
            "title": "기타 특이사항",
            "questions": [
                "특정 직원끼리 같은 날 OFF 금지/필수 조건이 있나요?",
                "연속 OFF (2일 연속) 가능한가요?",
                "근무표 완성 후 제출 대상/방식은? (예: 쿠팡 본사 이메일 송부)",
            ]
        },
    ]

    for sec in sections:
        # 섹션 제목
        merge(ws, row, 1, row, 5,
              f"  {sec['no']}.  {sec['title']}",
              font=normal_font(bold=True, color="1F4E79"),
              fill_color=COLOR_SECTION, align="left")
        ws.row_dimensions[row].height = 22
        row += 1

        for q in sec["questions"]:
            set_cell(ws, row, 1, "☐", font=normal_font(size=12), fill_color=COLOR_TODO, align="center")
            set_cell(ws, row, 2, q, font=normal_font(), fill_color=COLOR_TODO)
            set_cell(ws, row, 3, "", font=normal_font(), fill_color=COLOR_ANSWER)
            set_cell(ws, row, 4, "", font=normal_font(), fill_color=COLOR_ANSWER, align="center")
            set_cell(ws, row, 5, "", font=normal_font(), fill_color=COLOR_ANSWER, align="center")
            # 답변칸 높이 (줄바꿈 있으면 더 높게)
            ws.row_dimensions[row].height = 42 if "\n" in q else 28
            row += 1

        row += 1  # 섹션 간 간격

    # ── 요약표 ────────────────────────────────────────────
    merge(ws, row, 1, row, 5,
          "📋  확인 결과 요약",
          font=normal_font(bold=True, color="1F4E79"),
          fill_color=COLOR_SECTION, align="left")
    ws.row_dimensions[row].height = 22
    row += 1

    sum_headers = ["번호", "항목", "확인 결과 (요약)", "확인자", "확인 일자"]
    for col, h in enumerate(sum_headers, 1):
        set_cell(ws, row, col, h,
                 font=header_font(white=True),
                 fill_color="4472C4", align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    summary_items = [
        "동시 OFF 제한", "월 OFF 횟수 기준", "최소 유지 인원",
        "파출 투입 조건", "조리사 스케줄", "OFF 신청 우선순위",
        "연차 처리", "기타 특이사항",
    ]
    for i, item in enumerate(summary_items, 1):
        set_cell(ws, row, 1, i, font=normal_font(), fill_color=COLOR_SUMMARY, align="center")
        set_cell(ws, row, 2, item, font=normal_font(), fill_color=COLOR_SUMMARY)
        set_cell(ws, row, 3, "", font=normal_font(), fill_color=COLOR_ANSWER)
        set_cell(ws, row, 4, "", font=normal_font(), fill_color=COLOR_ANSWER, align="center")
        set_cell(ws, row, 5, "", font=normal_font(), fill_color=COLOR_ANSWER, align="center")
        ws.row_dimensions[row].height = 22
        row += 1

    # 시트 기본 설정
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"  # 헤더 고정

    wb.save(OUTPUT_PATH)
    print(f"파일 생성 완료: {OUTPUT_PATH}")

if __name__ == "__main__":
    build_checklist()
